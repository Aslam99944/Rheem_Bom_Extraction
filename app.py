"""
BOM Extraction — Production-Ready Hybrid Pipeline
Architecture: Azure Document Intelligence (Layout) + GPT-4o Vision (Text + Images)

Flow:
  1. PDF bytes → Azure Document Intelligence (prebuilt-layout) → structured text
  2. PDF bytes → PyMuPDF → page images (base64 PNG)
  3. Structured text + page images → GPT-4o Vision → JSON BOM data
  4. Validation & Dedup → Excel export
"""

import os
import io
import base64
import json
import time
import uuid
import logging
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from openai import AzureOpenAI
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF — PDF to image conversion

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Load .env and Configuration
# ---------------------------------------------------------------------------

load_dotenv()

# Azure OpenAI (GPT-4o for intelligent field mapping)
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "")
AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY", "")
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")

# Azure Document Intelligence (Layout model for PDF reading)
AZURE_DI_ENDPOINT = os.getenv("AZURE_DI_ENDPOINT", "")
AZURE_DI_KEY = os.getenv("AZURE_DI_KEY", "")

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class BOMItem(BaseModel):
    item: Optional[str] = ""
    part_number: Optional[str] = ""
    manufacturer: Optional[str] = ""
    description: Optional[str] = ""
    qty: Optional[str] = ""
    uom: Optional[str] = ""
    commodity: Optional[str] = ""
    type: Optional[str] = ""
    notes: Optional[str] = ""


class BOMExtractionResult(BaseModel):
    bom_items: list[BOMItem]

# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------

app = FastAPI(title="BOM Extraction POC")
app.mount("/static", StaticFiles(directory="static"), name="static")

# ---------------------------------------------------------------------------
# Azure Document Intelligence — Extract structured text from PDF
# ---------------------------------------------------------------------------

def analyze_document(file_bytes: bytes, content_type: str) -> dict:
    """
    Send PDF/image bytes to Azure Document Intelligence Layout model.
    Returns structured content: tables, paragraphs, and page-level text.
    """
    if not AZURE_DI_ENDPOINT or not AZURE_DI_KEY:
        raise HTTPException(
            status_code=500,
            detail="Azure Document Intelligence credentials not set. Check AZURE_DI_ENDPOINT and AZURE_DI_KEY in .env"
        )

    client = DocumentIntelligenceClient(
        endpoint=AZURE_DI_ENDPOINT,
        credential=AzureKeyCredential(AZURE_DI_KEY),
    )

    # Analyze using prebuilt-layout model (detects tables, text, structure)
    poller = client.begin_analyze_document(
        model_id="prebuilt-layout",
        body=file_bytes,
        content_type="application/octet-stream",
    )

    result = poller.result()
    logger.info(f"Azure DI: analyzed {len(result.pages)} page(s)")

    return result


# ---------------------------------------------------------------------------
# PDF to Base64 images (for GPT-4o Vision)
# ---------------------------------------------------------------------------

def pdf_pages_to_base64(file_bytes: bytes, dpi: int = 200) -> list[str]:
    """
    Convert all pages of a PDF to base64-encoded PNG images.
    Uses PyMuPDF (fitz) - no external dependencies like Poppler.
    Returns a list of base64 strings, one per page.
    """
    images = []
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc[page_num]
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            b64 = base64.b64encode(img_bytes).decode("utf-8")
            images.append(b64)
            logger.info(f"Page {page_num + 1}: converted to image ({pix.width}x{pix.height})")
        doc.close()
    except Exception as e:
        logger.warning(f"PDF to image conversion failed: {e}. Falling back to text-only.")
    return images


def build_structured_text(di_result) -> str:
    """
    Convert Azure Document Intelligence result into clean structured text
    that GPT-4o can understand without needing any images.
    """
    sections = []

    # --- Extract all tables with their cell data ---
    if di_result.tables:
        for t_idx, table in enumerate(di_result.tables):
            page_num = table.bounding_regions[0].page_number if table.bounding_regions else "?"

            # Build a matrix from cell data
            max_row = max(c.row_index for c in table.cells) + 1
            max_col = max(c.column_index for c in table.cells) + 1
            matrix = [["" for _ in range(max_col)] for _ in range(max_row)]

            for cell in table.cells:
                content = cell.content.strip() if cell.content else ""
                matrix[cell.row_index][cell.column_index] = content

            # Format as readable text table
            table_text = f"\n===== TABLE {t_idx + 1} (Page {page_num}, {max_row} rows x {max_col} cols) =====\n"
            for r_idx, row in enumerate(matrix):
                row_label = "HEADER" if r_idx == 0 else f"Row {r_idx}"
                table_text += f"  [{row_label}] " + " | ".join(row) + "\n"

            sections.append(table_text)
            logger.info(f"Table {t_idx + 1}: Page {page_num}, {max_row} rows x {max_col} cols")

    # --- Extract all paragraphs / free text (notes, callouts, annotations) ---
    if di_result.paragraphs:
        page_paragraphs = {}
        for para in di_result.paragraphs:
            page_num = para.bounding_regions[0].page_number if para.bounding_regions else 0
            if page_num not in page_paragraphs:
                page_paragraphs[page_num] = []
            content = para.content.strip() if para.content else ""
            if content:
                role = para.role if para.role else "text"
                page_paragraphs[page_num].append(f"  [{role}] {content}")

        for page_num in sorted(page_paragraphs.keys()):
            para_text = f"\n===== TEXT CONTENT (Page {page_num}) =====\n"
            para_text += "\n".join(page_paragraphs[page_num])
            sections.append(para_text)

    structured_text = "\n".join(sections)
    logger.info(f"Built structured text: {len(structured_text)} chars, {len(di_result.tables or [])} tables")
    return structured_text


# ---------------------------------------------------------------------------
# Fallback: Extract drawing info from structured text
# ---------------------------------------------------------------------------

def extract_drawing_info_from_text(text: str) -> dict:
    """
    Fallback: extract Drawing No. and Name from structured text
    if GPT-4o didn't return them in the JSON.
    """
    import re
    drawing_no = ""
    drawing_name = ""

    # Pattern: "Drawing No." or "Drawing No" or "DWG NO" followed by a value
    match = re.search(r'(?:Drawing\s*No\.?|DWG\s*NO\.?|Drawing\s*Number)\s*[:\s]*([A-Z0-9\-]+)', text, re.IGNORECASE)
    if match:
        drawing_no = match.group(1).strip()

    # Pattern: The drawing name usually appears near the drawing number
    # Look for title-like text near drawing number or in "TITLE" or "NAME" fields
    name_match = re.search(r'(?:TITLE|NAME|DESCRIPTION)\s*[:\s]*([A-Z][A-Z0-9\s/\-&\.]+)', text, re.IGNORECASE)
    if name_match:
        drawing_name = name_match.group(1).strip()

    # Also try: drawing number followed by the name on the same line
    if drawing_no and not drawing_name:
        line_match = re.search(rf'{re.escape(drawing_no)}\s+(.+?)(?:\n|$)', text)
        if line_match:
            candidate = line_match.group(1).strip()
            # Only use if it looks like a name (has letters, not just numbers)
            if len(candidate) > 5 and any(c.isalpha() for c in candidate):
                drawing_name = candidate

    if drawing_no or drawing_name:
        logger.info(f"Fallback drawing info: No={drawing_no}, Name={drawing_name}")

    return {"drawing_no": drawing_no, "drawing_name": drawing_name}


# ---------------------------------------------------------------------------
# GPT-4o Extraction Prompt — Single unified prompt (text-only, no images)
# ---------------------------------------------------------------------------

def build_extraction_prompt(structured_text: str) -> str:
    """
    Production-ready prompt that handles ALL BOM formats dynamically.
    Azure DI provides structured text; GPT-4o maps to BOM schema.
    """
    return f"""You are a Bill of Materials (BOM) extraction expert for industrial and engineering documents.

You are given TWO inputs:
1. **STRUCTURED TEXT** extracted by Azure Document Intelligence — contains accurate table data, text content, and annotations.
2. **PAGE IMAGES** of the actual document — use these to understand spatial layout, wire routing, dimensional annotations, and callout positions.

USE BOTH INPUTS TOGETHER:
- Use the TEXT for accurate part numbers, table cell values, and notes (text is more reliable for exact characters).
- Use the IMAGES for spatial understanding — which dimensions belong to which wires, how wires are routed, which callout labels refer to which component.
- If a value appears in text AND image, prefer the TEXT version (more accurate OCR).
- If something is visible in the IMAGE but missing from TEXT, extract it from the image.

===== DOCUMENT CONTENT =====
{structured_text}

===== YOUR TASK =====
Extract ALL BOM-relevant items from the document. Items come from:

1. **BOM / PARTS LIST TABLES**: Tables with columns like Item, Part Number, MFR, Description, Qty.
   → Extract EVERY data row. Count the rows in the table and verify your output has the same count.

2. **WIRE INFORMATION / CIRCUIT END TABLES**: Tables with wire-related columns (Wire Color, Wire Size, Wire Length, Connector Part Number, Terminal Part Number, etc.)
   → Each unique wire (unique combination of color + gauge) = one BOM item.

3. **ENGINEERING DRAWING CONTENT**: If NO formal BOM table exists, identify parts from callouts, labels, and annotations in the drawing.
   → Callout labels follow patterns like: "DESCRIPTION MFR P/N XXXXX" or "DESCRIPTION MFR XXXXX OR EQUIVALENT"
      Example: "PLUG HOUSING AMP P/N 1-480763-0" → part_number="1-480763-0", manufacturer="AMP", description="PLUG HOUSING"
      Example: "SOCKET CONTACT AMP P/N 350537-1" → part_number="350537-1", manufacturer="AMP", description="SOCKET CONTACT"
      Example: "RING TERMINAL AMP P/N 40595 OR EQUIVALENT" → part_number="40595", manufacturer="AMP", description="RING TERMINAL"
   → Wire labels in diagrams (e.g., '(BK) 19" BLACK', '(WH) 19" WHITE') are wire items. Capture EVERY instance even if they appear identical.
   → Wire LENGTHS come from dimensional annotations next to wires (e.g., '28 1/4"', '22 3/4"', '134"'). Extract these as the wire qty.
   → **WIRE HARNESS ROUTING**: In wire harness diagrams, multiple dimensional annotations represent SEGMENTS of the total wire path. ALL wires that run through the harness pass through ALL segments. You MUST:
     1. Identify ALL dimensional annotations in the diagram (e.g., 24.00, 18.00, 2.00 TYP).
     2. Trace each wire's path from one end to the other, through MIDWAY points.
     3. SUM all segment dimensions that each wire passes through to get the TOTAL wire length.
     Example: Diagram shows 24.00 (top segment) and 18.00 (bottom segment). All 3 wires (BK, WH, RD) run through both → each wire qty = 24 + 18 = 42.00
   → Part QUANTITIES come from annotations like "X REQ'D" (e.g., "6 REQ'D" means qty=6) or from counting instances in the diagram.
   → If no quantity is explicitly stated for a discrete part, infer from context (pin count, connector size, etc.) or set qty="1".

===== BOM TABLE EXTRACTION RULES =====

- Extract EVERY row from the BOM/Parts List table. Do NOT skip any row.
- "item": Exactly as shown in the table's Item column
- "part_number": Copy exactly from the Part Number column (e.g., "420C2PM12FL0", "430251200", "2-34146-1")
- "manufacturer": Copy from MFR column. Use full name when recognizable:
    - "OST" → "OST"
    - "MOLEX" → "Molex"
    - "TE" → "TE Connectivity"
- "description": ALL DESCRIPTIONS MUST BE UPPERCASE. Copy from Description column and convert to uppercase.
    Examples: "TWIST LOCK GROMMET", "GROMMET", "CABLE TIE", "RING TERMINAL", "12-PIN CONNECTOR"
- Special rules:
    - If a "cable tie" or "wire tie" item has a part number, extract the part number to the part_number field. The description stays as "CABLE TIE" or "WIRE TIE".
    - If description mentions only "label" with label text, set description to "LABEL".
- "qty": Copy from QTY column
- "uom": "EA" for discrete parts
- "commodity": "BOP" for purchased parts
- "type": Be specific:
    - If description mentions "PIN" or "Connector" → "Connector Housing"
    - If description mentions "Terminal" → "Terminal"
    - If description mentions "Label" → "Label"

===== WIRE TABLE EXTRACTION RULES =====

For each unique wire from Wire Information / Circuit End tables:

- "item": Number sequentially continuing from BOM table items
- "part_number": LEAVE EMPTY. Do NOT use Terminal Part Numbers or Connector Part Numbers as the wire's part number. Those belong in notes.
- "manufacturer": Leave empty for wires
- "description": Use format "{{WIRE_SIZE}}AWG,{{FULL_COLOR_NAME}} WIRE" — ALL UPPERCASE, comma after AWG, "WIRE" at the end.
    Convert standard wire color codes to full names:
    OG → ORANGE, YE → YELLOW, BK → BLACK, RD → RED, GN → GREEN,
    BU → BLUE, WH → WHITE, VT → VIOLET, GY → GRAY, PK → PINK, BN → BROWN
    R/Y or RD/YE → RED/YELLOW, BK/Y or BK/YE → BLACK/YELLOW
    Examples: "18AWG,RED WIRE", "10AWG,ORANGE WIRE", "20AWG,BLACK/YELLOW WIRE", "18AWG,RED/YELLOW WIRE"
- **AGGREGATION**: If multiple rows in the wire table describe the SAME wire type (same gauge and color), aggregate them into a SINGLE BOM item and calculate the **TOTAL sum** of their lengths.
    - Example: Row 1 = 18 R/Y 134", Row 2 = 18 R/Y 16". Output = One item, Description: "18AWG,RED/YELLOW WIRE", Qty: "150".
- "qty": Total summed length for this wire type. If the document NOTES mention a wire length tolerance (e.g., "WIRE LENGTH TOLERANCE ± .250" or "TOLERANCE ON FINISHED CABLE LENGTHS..."), append it to the sum: e.g., "176\"±0.250\""
- "uom": "Inch"
- "commodity": "Make"
- "type": "Wire & Cable"
- "notes": List ALL connector pin mappings for this wire type, separated by semicolons (e.g., "C1 Pin 1 to C2 Pin 1; C3 Pin 1 to C4 Pin 1"). Also include Terminal Part Numbers here.

===== DESCRIPTION CLEANING & EMBEDDED PART NUMBERS =====

- **Part Number Extraction**: Many documents embed the part number inside the description string (e.g., "TERMINAL RING TERMINAL 2-34149-1", "3X Female Quick Connect TE 3-520140-2"). 
    - You MUST scan the description for these patterns (e.g., "X-XXXXXX-X", "XXXXXX-X", "X-XXXX-X") and extract them to the "part_number" field.
- **Strict Cleaning**: Once a part number or manufacturer name (like "TE", "TE CONNECTIVITY", "MOLEX") has been extracted to its respective column, you MUST REMOVE IT from the "description" field.
    - Example Wrong: description="3X Female Quick Connect TE 3-520140-2", part_number="3-520140-2", manufacturer="TE Connectivity"
    - Example Correct: description="3X Female Quick Connect", part_number="3-520140-2", manufacturer="TE Connectivity"
- The "description" column should ONLY contain the human-readable description of the component in UPPERCASE (e.g., "3X FEMALE QUICK CONNECT", "RING TERMINAL", "WIRE TIES").

===== DO NOT EXTRACT AS BOM ITEMS =====

These are NOT BOM items — do NOT create entries for them:

1. **Drawing numbers / titles from the title block** (e.g., "AP22971 - WIRE HARNESS - CTA HARNESS W/POWER" is the drawing title, NOT a part)
2. **Alternate parts mentioned in NOTES** (e.g., "2-34149-1 CAN BE USED AS AN ALTERNATE..." — this is a note about an alternative, NOT a BOM item. Put it as a note on the relevant item instead.)
3. **Tolerance specifications** and tolerance tables
4. **Assembly instructions** (e.g., "Tape each connector...", "Apply heat shrink...")
5. **Manufacturing standards** (e.g., "IPC-A-620 CLASS 2", "UL 1569")
6. **Revision history**, dates, signatures, approval stamps
7. **The wire harness assembly itself** — the drawing DESCRIBES it, it is not a separate BOM item
8. **Assembly-level part numbers from notes** — e.g., "RHEEM PART # 45-24258-02", "ASSEMBLY P/N XXXXX". These identify the ASSEMBLY being built, NOT a component to be purchased. Do NOT create a BOM item for them.

===== QUANTITY FORMAT =====

- All quantities MUST be in DECIMAL format. Convert fractions to decimals:
    - 28 1/4 → 28.25
    - 22 3/4 → 22.75
    - 28 1/2 → 28.5
    - 6 1/8 → 6.125
- If no quantity is found for a discrete part, set qty="1".

===== NOTES FIELD GUIDANCE =====

- For BOM table items: Leave notes empty unless there is specific consumption or usage info. Do NOT write "Details not clear from the Drawing" if the part number and description are already clear.
- For wire items: Include connector-pin-to-connector-pin mapping from wire table (e.g., "C1 Pin 1 to C2 Pin 1")
- If a note in the document mentions an alternate part for an existing item, add it as a note on THAT item (e.g., "Alternate: 2-34149-1 Ring Terminal")

===== DRAWING METADATA =====
Also extract the drawing metadata from the title block:
- "drawing_no": The drawing number (e.g., "651582", "AP22971")
- "drawing_name": The drawing title/name (e.g., "HARNESS ASSY-GAS VALVE/ SAFETY BD/AIR SW", "WIRE HARNESS - CTA HARNESS W/POWER")

===== OUTPUT FORMAT =====
Return a JSON object with TWO keys:
{{
  "drawing_info": {{"drawing_no": "...", "drawing_name": "..."}},
  "bom_items": [...]
}}
Each bom_item: {{"item", "part_number", "manufacturer", "description", "qty", "uom", "commodity", "type", "notes"}}
All values MUST be strings. Empty fields = "".
Return ONLY valid JSON. No commentary, no markdown."""


# ---------------------------------------------------------------------------
# Post-Extraction Validation & Cleanup
# ---------------------------------------------------------------------------

def validate_bom_items(items: list[dict]) -> list[dict]:
    """Clean, deduplicate, and validate extracted BOM items."""
    cleaned = []
    seen = set()

    for item in items:
        # Skip items with empty description AND empty part_number (true ghost rows)
        desc = str(item.get("description", "")).strip()
        part_num = str(item.get("part_number", "")).strip()
        if not desc and not part_num:
            logger.info(f"Skipping ghost row: {item}")
            continue

        # Normalize whitespace in all fields
        for key in item:
            if isinstance(item[key], str):
                item[key] = " ".join(item[key].split()).strip()

        # Convert fractional qty to decimal (e.g., "28 1/4" → "28.25", "3/4" → "0.75")
        qty_val = str(item.get("qty", "")).strip()
        if qty_val:
            import re
            from fractions import Fraction
            frac_match = re.match(r'^(\d+)\s+(\d+/\d+)(.*)$', qty_val)
            if frac_match:
                whole = int(frac_match.group(1))
                frac = float(Fraction(frac_match.group(2)))
                suffix = frac_match.group(3).strip()
                decimal_val = whole + frac
                qty_str = f"{decimal_val:g}"
                item["qty"] = f"{qty_str} {suffix}".strip() if suffix else qty_str
            else:
                standalone_frac = re.match(r'^(\d+/\d+)(.*)$', qty_val)
                if standalone_frac:
                    frac = float(Fraction(standalone_frac.group(1)))
                    suffix = standalone_frac.group(2).strip()
                    qty_str = f"{frac:g}"
                    item["qty"] = f"{qty_str} {suffix}".strip() if suffix else qty_str
        # Robustness: Remove extracted part_num and manufacturer from desc
        if part_num and part_num.lower() in desc.lower():
            import re
            item["description"] = re.sub(re.escape(part_num), "", desc, flags=re.IGNORECASE).strip()
            desc = item["description"]
        
        mfr = str(item.get("manufacturer", "")).strip()
        if mfr and mfr.lower() in desc.lower():
            import re
            item["description"] = re.sub(re.escape(mfr), "", desc, flags=re.IGNORECASE).strip()
            # Handle common variations like "TE Connectivity" part vs just "TE"
            if mfr.lower() == "te connectivity":
                item["description"] = re.sub(r'\bTE\b', "", item["description"], flags=re.IGNORECASE).strip()
            desc = item["description"]

        # Final cleanup of double spaces or separators like ":" or "-" at start/end
        item["description"] = re.sub(r'^\s*[:\-]\s*', '', item["description"])
        item["description"] = re.sub(r'\s*[:\-]\s*$', '', item["description"])
        item["description"] = " ".join(item["description"].split())
        desc = item["description"]

        # Force all descriptions to UPPERCASE
        item["description"] = item["description"].upper()
        desc = item["description"]
        # Dedup key (description + qty) — skip for wires, they get summed later
        if item.get("type") != "Wire & Cable":
            dedup_key = (desc.lower(), item.get("qty", "").strip())
            if dedup_key in seen:
                logger.info(f"Removing duplicate: {desc}")
                continue
            seen.add(dedup_key)

        cleaned.append(item)

    # Sort by item number (handle non-numeric gracefully)
    def sort_key(item):
        try:
            return (0, int(str(item.get("item", "0")).strip()))
        except ValueError:
            return (1, 0)  # Non-numeric items go to the end

    cleaned.sort(key=sort_key)

    # --- Wire Aggregation (Sum Qty for same AWG/Color) ---
    final_items = []
    wire_map = {} # desc -> index in final_items

    for item in cleaned:
        if item.get("type") == "Wire & Cable":
            desc = item.get("description", "").strip()
            if desc in wire_map:
                idx = wire_map[desc]
                existing = final_items[idx]
                
                # Sum the quantities
                try:
                    import re
                    # Extract numeric part (handles integers, floats, and trailing units/tolerances)
                    def extract_num(s):
                        match = re.search(r"(\d+\.?\d*)", str(s))
                        return float(match.group(1)) if match else 0.0
                    
                    new_qty = extract_num(existing["qty"]) + extract_num(item["qty"])
                    
                    # Keep original suffix (e.g. "±0.250") if present
                    suffix_match = re.search(r"([^\d.].*)$", str(existing["qty"]))
                    suffix = suffix_match.group(1) if suffix_match else ""
                    
                    # Format as integer if no decimal, else float
                    qty_str = f"{int(new_qty)}" if new_qty.is_integer() else f"{new_qty:.2f}"
                    existing["qty"] = f"{qty_str}{suffix}"
                    
                    # Merge notes
                    if item.get("notes"):
                        existing["notes"] = (existing.get("notes", "") + "; " + item["notes"]).strip("; ")
                except Exception as e:
                    logger.warning(f"Failed to sum wire qty for {desc}: {e}")
                    final_items.append(item) # fallback to separate items
            else:
                wire_map[desc] = len(final_items)
                final_items.append(item)
        else:
            final_items.append(item)

    logger.info(f"Validation: {len(items)} raw -> {len(final_items)} cleaned (sorted & wire-summed)")
    return final_items


# ---------------------------------------------------------------------------
# Main Extraction Pipeline
# ---------------------------------------------------------------------------

def extract_bom(file_bytes: bytes, content_type: str) -> dict:
    """
    Production pipeline:
    1. Azure Document Intelligence → structured text
    2. GPT-4o (text-only) → BOM JSON
    3. Validation → clean data
    """
    # --- Step 1: Azure Document Intelligence ---
    di_result = analyze_document(file_bytes, content_type)
    structured_text = build_structured_text(di_result)

    if not structured_text.strip():
        logger.warning("Azure DI returned empty content")
        return {
            "bom_data": {"bom_items": []},
            "token_info": {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0, "estimated_cost_usd": 0},
            "structured_text": "",
            "pages_analyzed": len(di_result.pages) if di_result.pages else 0,
        }

    # --- Step 2: GPT-4o Text-only extraction ---
    if not AZURE_OPENAI_API_KEY or not AZURE_OPENAI_ENDPOINT:
        raise HTTPException(
            status_code=500,
            detail="Azure OpenAI credentials not set. Check .env"
        )

    openai_client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version=AZURE_OPENAI_API_VERSION,
    )

    prompt = build_extraction_prompt(structured_text)

    # --- Step 2b: Convert PDF pages to images ---
    page_images = []
    if content_type == "application/pdf":
        page_images = pdf_pages_to_base64(file_bytes)
        logger.info(f"Converted {len(page_images)} PDF pages to images for Vision")
    else:
        # For images (PNG/JPG/TIFF), encode the file directly
        b64 = base64.b64encode(file_bytes).decode("utf-8")
        page_images = [b64]
        logger.info("Encoded uploaded image for Vision")

    # --- Step 2c: Build GPT-4o Vision message ---
    content_parts = [
        {"type": "text", "text": prompt}
    ]
    for idx, img_b64 in enumerate(page_images):
        mime = "image/png"
        content_parts.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:{mime};base64,{img_b64}",
                "detail": "high"
            }
        })

    response = openai_client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[{"role": "user", "content": content_parts}],
        temperature=0.0,
        max_tokens=8192,
        response_format={"type": "json_object"},
    )

    raw_text = response.choices[0].message.content or "{}"
    bom_data = json.loads(raw_text)
    logger.info(f"GPT-4o response keys: {list(bom_data.keys())}")
    logger.info(f"GPT-4o raw response (first 500 chars): {raw_text[:500]}")

    items = bom_data.get("bom_items", [])
    drawing_info = bom_data.get("drawing_info", {})

    # Fallback: if GPT-4o didn't return drawing_info, try to extract from structured text
    if not drawing_info or (not drawing_info.get("drawing_no") and not drawing_info.get("drawing_name")):
        drawing_info = extract_drawing_info_from_text(structured_text)

    logger.info(f"GPT-4o extracted {len(items)} raw items, drawing: {drawing_info}")

    # --- Step 3: Validate ---
    items = validate_bom_items(items)

    # Token accounting
    token_info = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
    if response.usage:
        token_info["input_tokens"] = response.usage.prompt_tokens
        token_info["output_tokens"] = response.usage.completion_tokens
        token_info["total_tokens"] = response.usage.total_tokens

    input_cost = (token_info["input_tokens"] / 1_000_000) * 2.50
    output_cost = (token_info["output_tokens"] / 1_000_000) * 10.00
    token_info["estimated_cost_usd"] = round(input_cost + output_cost, 6)

    return {
        "bom_data": {"bom_items": items},
        "drawing_info": drawing_info,
        "token_info": token_info,
        "structured_text": structured_text,
        "pages_analyzed": len(di_result.pages) if di_result.pages else 0,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def create_excel(bom_items: list[dict], filename: str, drawing_info: dict = None) -> Path:
    """Create a styled Excel file with drawing metadata header + BOM data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Data"

    # --- Styles ---
    meta_label_font = Font(name="Calibri", bold=True, size=11, color="000000")
    meta_value_font = Font(name="Calibri", size=11, color="000000")
    meta_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow like the PDF
    meta_align = Alignment(vertical="center")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    current_row = 1

    # --- Drawing Metadata Header (Row 1) ---
    if drawing_info and (drawing_info.get("drawing_no") or drawing_info.get("drawing_name")):
        drawing_no = drawing_info.get("drawing_no", "")
        drawing_name = drawing_info.get("drawing_name", "")

        # "Drawing No." label
        cell = ws.cell(row=current_row, column=1, value="Drawing No.")
        cell.font = meta_label_font
        cell.fill = meta_fill
        cell.alignment = meta_align
        cell.border = thin_border

        # Drawing number value
        cell = ws.cell(row=current_row, column=2, value=drawing_no)
        cell.font = meta_value_font
        cell.fill = meta_fill
        cell.alignment = meta_align
        cell.border = thin_border

        # Drawing name (merged across remaining columns)
        cell = ws.cell(row=current_row, column=3, value=drawing_name)
        cell.font = meta_value_font
        cell.fill = meta_fill
        cell.alignment = meta_align
        cell.border = thin_border
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=9)

        current_row += 1  # Move to next row

    # --- BOM Header Row ---
    headers = ["Item", "Part Number", "Manufacturer", "Description", "Qty", "UOM", "Commodity", "Type", "Notes"]
    field_keys = ["item", "part_number", "manufacturer", "description", "qty", "uom", "commodity", "type", "notes"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- BOM Data Rows ---
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    alt_fill = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")

    for row_idx, item in enumerate(bom_items, current_row + 1):
        for col_idx, key in enumerate(field_keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # --- Column widths ---
    col_widths = [8, 18, 22, 40, 12, 8, 16, 20, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze below the header row (accounts for metadata row)
    ws.freeze_panes = f"A{current_row + 1}"

    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def serve_ui():
    return HTMLResponse(content=Path("static/index.html").read_text(encoding="utf-8"))


@app.post("/upload")
async def upload_and_extract(file: UploadFile = File(...)):
    """Upload BOM PDF/image -> Azure DI + GPT-4o -> JSON + Excel."""
    start_time = time.time()

    # Validate file type
    ext = Path(file.filename or "").suffix.lower()
    content_type = file.content_type or ""
    allowed_ext = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif"}

    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")

    file_bytes = await file.read()

    logger.info(f"Processing: {file.filename} ({len(file_bytes)} bytes)")

    # Run the extraction pipeline
    result = extract_bom(file_bytes, content_type)

    processing_time = round(time.time() - start_time, 2)
    bom_items = result["bom_data"].get("bom_items", [])
    drawing_info = result.get("drawing_info", {})
    token_info = result["token_info"]

    # Generate Excel with drawing metadata — filename matches uploaded file
    original_stem = Path(file.filename or "bom_extract").stem
    excel_filename = f"{original_stem}.xlsx"
    create_excel(bom_items, excel_filename, drawing_info=drawing_info)

    return {
        "success": True,
        "bom_items": bom_items,
        "drawing_info": drawing_info,
        "ocr_text": result["structured_text"],
        "token_info": token_info,
        "processing_time_seconds": processing_time,
        "row_count": len(bom_items),
        "excel_filename": excel_filename,
    }


@app.get("/download/{filename}")
async def download_excel(filename: str):
    filepath = OUTPUT_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(
        path=filepath, filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    print("\n[*] BOM Extraction POC (Azure Document Intelligence + GPT-4o)")
    print("[*] No Tesseract, No VLM, No Image Processing")
    print("[*] Running at http://localhost:8000\n")
    uvicorn.run(app, host="0.0.0.0", port=8000)
